# -*- coding: utf-8 -*-
import pandas as pd
import mysql.connector
import os
import glob
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# New folder
month = 'Январь_Февраль'
archive = os.mkdir(f'path_\\Archive_{month}')

# Connect to the MySQL database
cnx = mysql.connector.connect(user='user', password='pass',
                              host='host',
                              database='db')
cursor = cnx.cursor()

# Профили
queries = {
'COVID-19' :"SELECT Data_Analyst.zayav_mkb.Код_услуги, Data_Analyst.mes.Наименование, SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2019 ВСЕГО', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2019 Экстренная абс', (SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)) / SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2019 Экстренная %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2019 СМП', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2019 Самотек', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2019 Плановая абс', (SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)) / SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2019 Плановая %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2019 План', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2019 Военкомат', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.zayav_mkb.Признак_СКП_не_СКП = 'ДА' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS 'СКП абс', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 AND Data_Analyst.zayav_mkb.Признак_СКП_не_СКП = 'ДА' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)/SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS 'СКП %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2020 ВСЕГО', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2020 Экстренная абс', (SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)) / SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2020 Экстренная %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2020 СМП', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2020 Самотек', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2020 Плановая абс', (SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)) / SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2020 Плановая %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2020 План', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2020 Военкомат', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.zayav_mkb.Признак_СКП_не_СКП = 'ДА' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS 'СКП абс', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2020 AND Data_Analyst.zayav_mkb.Признак_СКП_не_СКП = 'ДА' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)/SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS 'СКП %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2021 ВСЕГО', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2021 Экстренная абс', (SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)) / SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2021 Экстренная %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2021 СМП', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2021 Самотек', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2021 Плановая абс', (SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)) / SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2021 Плановая %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2021 План', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2021 Военкомат', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.zayav_mkb.Признак_СКП_не_СКП = 'ДА' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS 'СКП абс', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2021 AND Data_Analyst.zayav_mkb.Признак_СКП_не_СКП = 'ДА' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)/SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS 'СКП %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2022 ВСЕГО', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2022 Экстренная абс', (SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)) / SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2022 Экстренная %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2022 СМП', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2022 Самотек', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2022 Плановая абс', (SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)) / SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2022 Плановая %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2022 План', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2022 Военкомат', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.zayav_mkb.Признак_СКП_не_СКП = 'ДА' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS 'СКП абс', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2022 AND Data_Analyst.zayav_mkb.Признак_СКП_не_СКП = 'ДА' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)/SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS 'СКП %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2023 ВСЕГО', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2023 Экстренная абс', (SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)) / SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2023 Экстренная %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.kanal.Канал = 'смп' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2023 СМП', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.kanal.Канал = 'самотек' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2023 Самотек', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2023 Плановая абс', (SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)+SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)) / SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) as '2023 Плановая %', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.kanal.Канал = 'план' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2023 План', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.kanal.Канал = 'военкомат' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS '2023 Военкомат', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.zayav_mkb.Признак_СКП_не_СКП = 'ДА' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS 'СКП абс', SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 AND Data_Analyst.zayav_mkb.Признак_СКП_не_СКП = 'ДА' THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END)/SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2019 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) AS 'СКП %' FROM Data_Analyst.zayav_mkb LEFT JOIN Data_Analyst.kanal ON Data_Analyst.zayav_mkb.Канал_госпитализации = Data_Analyst.kanal.ID left join Data_Analyst.hosp on Data_Analyst.zayav_mkb.Профиль = Data_Analyst.hosp.Профиль_с_ошибками left join Data_Analyst.mo on Data_Analyst.zayav_mkb.ID = Data_Analyst.mo.ID left join Data_Analyst.mes on Data_Analyst.zayav_mkb.Код_услуги = Data_Analyst.mes.ID_МЭС left join Data_Analyst.mes_profile on Data_Analyst.zayav_mkb.Код_услуги = Data_Analyst.mes_profile.ID_МЭС where case when Data_Analyst.hosp.Профиль_2023 = 'Акушерство и гинекология'  then Data_Analyst.mes_profile.Профиль when Data_Analyst.zayav_mkb.Код_услуги = '84010' or Data_Analyst.zayav_mkb.Код_услуги = '84020' or Data_Analyst.zayav_mkb.Код_услуги = '84030' or Data_Analyst.zayav_mkb.Код_услуги = '184010' or Data_Analyst.zayav_mkb.Код_услуги = '184011' or Data_Analyst.zayav_mkb.Код_услуги = '184020' or Data_Analyst.zayav_mkb.Код_услуги = '184021' or Data_Analyst.zayav_mkb.Код_услуги = '184030' or Data_Analyst.zayav_mkb.Код_услуги = '184031' then Data_Analyst.hosp.Профиль_2023 when Data_Analyst.hosp.Профиль_2023 != 'Акушерство и гинекология' then Data_Analyst.hosp.Профиль_2023 end = 'COVID-19'  and Data_Analyst.zayav_mkb.ТИП_пациента = 'МСК' and Data_Analyst.mes.Тип_МЭС = 'СТАЦ' and Data_Analyst.mo.Ведомство = 'ДЗМ' and Data_Analyst.mo.В_Д = 'В' group by Data_Analyst.zayav_mkb.Код_услуги order by SUM(CASE WHEN substring(Data_Analyst.zayav_mkb.Период, -9, 4) = 2023 THEN Data_Analyst.zayav_mkb.Количество ELSE 0 END) DESC;",
            }

# Сохраняем ехель
for profile, query in queries.items():
    cursor.execute(query)
    results = cursor.fetchall()
    df = pd.DataFrame(results, columns=[i[0] for i in cursor.description])
    with pd.ExcelWriter(f'path_{month}\\{profile}.xlsx') as writer:
        df.to_excel(writer, sheet_name=profile, index=False)

# БД бай бай
cursor.close()
cnx.close()

month = 'Январь_Март'
folder_path = f'path_\\Archive_{month}'


for file_path in glob.glob(os.path.join(folder_path, '*.xlsx')):
    workbook = load_workbook(filename=file_path)

    worksheet = workbook.active

    for worksheet in workbook.worksheets:
        for row in worksheet.rows:
            for cell in row:
                cell.border = openpyxl.styles.Border()

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                try:
                    cell.value = float(cell.value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = 'General;-"General";'
                    cell.number_format = '# ##0'
                except (ValueError, TypeError):
                    pass
    font = Font(name='Arial Narrow')

    for row in worksheet.rows:
        for cell in row:
            cell.font = font

    worksheet.insert_rows(1)
    worksheet.insert_rows(1)
    worksheet.insert_rows(1)
    worksheet.insert_rows(1)
    worksheet.insert_rows(6)
    worksheet.insert_rows(6)
    worksheet.insert_rows(6)

    # Добавить колонки
    worksheet.insert_cols(0, amount=1)
    worksheet.insert_cols(3, amount=1)
    worksheet.column_dimensions['C'].width = 0.9
    worksheet.insert_cols(5, amount=1)
    worksheet.column_dimensions['E'].width = 0.9
    worksheet.insert_cols(7, amount=1)  # 2019 (%)
    worksheet.insert_cols(18, amount=1)
    worksheet.column_dimensions['R'].width = 0.9
    worksheet.insert_cols(20, amount=1)  # 2020 (%)
    worksheet.insert_cols(31, amount=1)
    worksheet.column_dimensions['AE'].width = 0.9
    worksheet.insert_cols(33, amount=1)  # 2021 (%)
    worksheet.insert_cols(44, amount=1)
    worksheet.column_dimensions['AR'].width = 0.9
    worksheet.insert_cols(46, amount=1)  # 2022 (%)
    worksheet.insert_cols(57, amount=1)
    worksheet.column_dimensions['BE'].width = 0.9
    worksheet.insert_cols(59, amount=1)  # 2023 (%)

    border_style = openpyxl.styles.Side(border_style='thin', color='D9D9D9')
    border = openpyxl.styles.Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border

    worksheet[
        'B1'] = 'Профиль ' + worksheet.title + ': 2019 г. - 2023 г. (по Март)'
    worksheet['B1'].font = openpyxl.styles.Font(name='Arial Narrow', size=18, bold=True, italic=True, color='0000FF')

    worksheet.column_dimensions.group('J', 'K', hidden=True)
    worksheet.column_dimensions.group('N', 'Q', hidden=True)
    worksheet.column_dimensions.group('W', 'X', hidden=True)
    worksheet.column_dimensions.group('AA', 'AD', hidden=True)
    worksheet.column_dimensions.group('AJ', 'AK', hidden=True)
    worksheet.column_dimensions.group('AN', 'AQ', hidden=True)
    worksheet.column_dimensions.group('AW', 'AX', hidden=True)
    worksheet.column_dimensions.group('BA', 'BD', hidden=True)
    worksheet.column_dimensions.group('BJ', 'BK', hidden=True)
    worksheet.column_dimensions.group('BN', 'BQ', hidden=True)

    # Года
    worksheet.merge_cells('F2:Q2')
    worksheet['F2'] = '2019 год'
    worksheet['F2'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['F2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('S2:AD2')
    worksheet['S2'] = '2020 год'
    worksheet['S2'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['S2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AF2:AQ2')
    worksheet['AF2'] = '2021 год'
    worksheet['AF2'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AF2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AS2:BD2')
    worksheet['AS2'] = '2022 год'
    worksheet['AS2'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AS2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('BF2:BQ2')
    worksheet['BF2'] = '2023 год (Январь - Март)'
    worksheet['BF2'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['BF2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Экстренная/Плановая 2019
    worksheet.merge_cells('F3:G4')
    worksheet['F3'] = 'ВСЕГО'
    worksheet['F3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['F3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('H3:K3')
    worksheet['H3'] = 'Экстренная'
    worksheet['H3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['H3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('L3:O3')
    worksheet['L3'] = 'Плановая'
    worksheet['L3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['L3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('P3:Q3')
    worksheet['P3'] = 'СКП'
    worksheet['P3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['P3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('H4:K4')
    worksheet['H4'] = 'Итого'
    worksheet['H4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['H4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('L4:O4')
    worksheet['L4'] = 'Итого'
    worksheet['L4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['L4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('P4:Q4')
    worksheet['P4'] = 'Итого'
    worksheet['P4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['P4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Экстренная/Плановая 2020
    worksheet.merge_cells('S3:T4')
    worksheet['S3'] = 'ВСЕГО'
    worksheet['S3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['S3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('U3:X3')
    worksheet['U3'] = 'Экстренная'
    worksheet['U3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['U3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('Y3:AB3')
    worksheet['Y3'] = 'Плановая'
    worksheet['Y3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['Y3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AC3:AD3')
    worksheet['AC3'] = 'СКП'
    worksheet['AC3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AC3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('U4:X4')
    worksheet['U4'] = 'Итого'
    worksheet['U4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['U4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('Y4:AB4')
    worksheet['Y4'] = 'Итого'
    worksheet['Y4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['Y4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AC4:AD4')
    worksheet['AC4'] = 'Итого'
    worksheet['AC4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AC4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Экстренная/Плановая 2021
    worksheet.merge_cells('AF3:AG4')
    worksheet['AF3'] = 'ВСЕГО'
    worksheet['AF3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AF3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AH3:AK3')
    worksheet['AH3'] = 'Экстренная'
    worksheet['AH3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AH3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AL3:AO3')
    worksheet['AL3'] = 'Плановая'
    worksheet['AL3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AL3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AP3:AQ3')
    worksheet['AP3'] = 'СКП'
    worksheet['AP3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AP3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AH4:AK4')
    worksheet['AH4'] = 'Итого'
    worksheet['AH4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AH4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AL4:AO4')
    worksheet['AL4'] = 'Итого'
    worksheet['AL4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AL4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AP4:AQ4')
    worksheet['AP4'] = 'Итого'
    worksheet['AP4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AP4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Экстренная/Плановая 2022
    worksheet.merge_cells('AS3:AT4')
    worksheet['AS3'] = 'ВСЕГО'
    worksheet['AS3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AS3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AU3:AX3')
    worksheet['AU3'] = 'Экстренная'
    worksheet['AU3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AU3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AY3:BB3')
    worksheet['AY3'] = 'Плановая'
    worksheet['AY3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AY3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('BC3:BD3')
    worksheet['BC3'] = 'СКП'
    worksheet['BC3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['BC3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AU4:AX4')
    worksheet['AU4'] = 'Итого'
    worksheet['AU4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AU4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('AY4:BB4')
    worksheet['AY4'] = 'Итого'
    worksheet['AY4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['AY4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('BC4:BD4')
    worksheet['BC4'] = 'Итого'
    worksheet['BC4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['BC4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Экстренная/Плановая 2023
    worksheet.merge_cells('BF3:BG4')
    worksheet['BF3'] = 'ВСЕГО'
    worksheet['BF3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['BF3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('BH3:BK3')
    worksheet['BH3'] = 'Экстренная'
    worksheet['BH3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['BH3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('BL3:BO3')
    worksheet['BL3'] = 'Плановая'
    worksheet['BL3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['BL3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('BP3:BQ3')
    worksheet['BP3'] = 'СКП'
    worksheet['BP3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['BP3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('BH4:BK4')
    worksheet['BH4'] = 'Итого'
    worksheet['BH4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['BH4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('BL4:BO4')
    worksheet['BL4'] = 'Итого'
    worksheet['BL4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['BL4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('BP4:BQ4')
    worksheet['BP4'] = 'Итого'
    worksheet['BP4'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['BP4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Код_услуги & Наименование & Номер
    worksheet.merge_cells('A3:A5')
    worksheet['A3'] = '№'
    worksheet['A3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['A3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('B3:B5')
    worksheet['B3'] = 'Код услуги'
    worksheet['B3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet['B3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.merge_cells('D3:D5')
    worksheet['D3'] = 'Наименование услуги'
    worksheet['D3'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['D3'].font = openpyxl.styles.Font(name='Arial Narrow', size=12, bold=True)
    worksheet.column_dimensions['D'].width = 42
    column_b = worksheet[get_column_letter(4)]
    for cell in column_b:
        cell.alignment = Alignment(wrap_text=True)
    column_f = worksheet[get_column_letter(6)]
    column_s = worksheet[get_column_letter(19)]
    column_af = worksheet[get_column_letter(32)]
    column_as = worksheet[get_column_letter(45)]
    column_bf = worksheet[get_column_letter(58)]
    for cell in column_f:
        cell.font = openpyxl.styles.Font(name='Arial Narrow', size=11, bold=True)
    for cell in column_s:
        cell.font = openpyxl.styles.Font(name='Arial Narrow', size=11, bold=True)
    for cell in column_af:
        cell.font = openpyxl.styles.Font(name='Arial Narrow', size=11, bold=True)
    for cell in column_as:
        cell.font = openpyxl.styles.Font(name='Arial Narrow', size=11, bold=True)
    for cell in column_bf:
        cell.font = openpyxl.styles.Font(name='Arial Narrow', size=11, bold=True)

    # Сумма итогов в строке 7
    worksheet.merge_cells('B7:D7')
    worksheet['B7'] = 'ИТОГО'
    worksheet['B7'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['B7'].font = openpyxl.styles.Font(name='Arial Narrow', size=16, bold=True)
    for column, target_cell in zip(
            ['F', 'H', 'J', 'K', 'L', 'N', 'O', 'P', 'S', 'U', 'W', 'X', 'Y', 'AA', 'AB', 'AC', 'AF', 'AH', 'AJ', 'AK',
             'AL', 'AN', 'AO', 'AP', 'AS', 'AU', 'AW', 'AX', 'AY', 'BA', 'BB', 'BC', 'BF', 'BH', 'BJ', 'BK',
             'BL', 'BN', 'BO', 'BP'],
            ['F7', 'H7', 'J7', 'K7', 'L7', 'N7', 'O7', 'P7', 'S7', 'U7', 'W7', 'X7', 'Y7', 'AA7', 'AB7', 'AC7', 'AF7',
             'AH7', 'AJ7', 'AK7', 'AL7', 'AN7', 'AO7', 'AP7', 'AS7', 'AU7', 'AW7', 'AX7', 'AY7', 'BA7', 'BB7', 'BC7',
             'BF7', 'BH7', 'BJ7', 'BK7', 'BL7', 'BN7', 'BO7', 'BP7']):
        formula = f"=SUM({column}9:{column}1000)"
        worksheet[target_cell] = formula
        worksheet[target_cell].font = openpyxl.styles.Font(name='Arial Narrow', size=14, bold=True)
        worksheet[target_cell].alignment = Alignment(horizontal='center', vertical='center')
        worksheet[target_cell].number_format = '# ##0'
    worksheet.merge_cells('F7:G7')  # Всего 2019
    worksheet.merge_cells('S7:T7')  # Всего 2020
    worksheet.merge_cells('AF7:AG7')  # Всего 2021
    worksheet.merge_cells('AS7:AT7')  # Всего 2022
    worksheet.merge_cells('BF7:BG7')  # Всего 2023

    # Переименование столбцов
    values = ['абс', '%', 'абс', '%', 'СМП', 'Самотек', 'абс', '%', 'План', 'Военкомат', 'абс', '%', 'абс', '%', 'абс',
              '%', 'СМП', 'Самотек', 'абс', '%', 'План', 'Военкомат', 'абс', '%', 'абс', '%', 'абс', '%', 'СМП',
              'Самотек', 'абс', '%', 'План', 'Военкомат', 'абс', '%', 'абс', '%', 'абс', '%', 'СМП', 'Самотек', 'абс',
              '%', 'План', 'Военкомат', 'абс', '%', 'абс', '%', 'абс', '%', 'СМП', 'Самотек', 'абс', '%', 'План',
              'Военкомат', 'абс', '%', ]
    columns = ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA',
               'AB', 'AC', 'AD', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AS', 'AT',
               'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL',
               'BM', 'BN', 'BO', 'BP', 'BQ']
    row_index = 5

    # define a dictionary of columns and their corresponding column numbers
    column_dict = {
        'G': 7,
        'I': 9,
        'M': 13,
        'Q': 17,
        'T': 20,
        'V': 22,
        'Z': 26,
        'AD': 30,
        'AG': 33,
        'AI': 35,
        'AM': 39,
        'AQ': 43,
        'AT': 46,
        'AV': 48,
        'AZ': 52,
        'BD': 56,
        'BG': 59,
        'BI': 61,
        'BM': 65,
        'BQ': 69,
    }

    for col, value in zip(columns, values):
        worksheet[f'{col}{row_index}'] = value

    for col in column_dict:
        column_num = column_dict[col]
        for row in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=column_num)
            value = cell.value
            cell.number_format = '0%'
            cell.font = openpyxl.styles.Font(name='Arial Narrow', size=11, italic=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # % от всего
    for row in range(9, worksheet.max_row + 1):
        formula = f'=IFERROR(F{row}/$F$7,"")'
        cell = worksheet.cell(row=row, column=7)
        cell.value = formula
    for row in range(9, worksheet.max_row + 1):
        formula = f'=IFERROR(S{row}/$S$7,"")'
        cell = worksheet.cell(row=row, column=20)
        cell.value = formula
    for row in range(9, worksheet.max_row + 1):
        formula = f'=IFERROR(AF{row}/$AF$7,"")'
        cell = worksheet.cell(row=row, column=33)
        cell.value = formula
    for row in range(9, worksheet.max_row + 1):
        formula = f'=IFERROR(AS{row}/$AS$7,"")'
        cell = worksheet.cell(row=row, column=46)
        cell.value = formula
    for row in range(9, worksheet.max_row + 1):
        formula = f'=IFERROR(BF{row}/$BF$7,"")'
        cell = worksheet.cell(row=row, column=59)
        cell.value = formula

    # % от всего (экстренная/плановая/СКП)
    # 2019
    worksheet['I7'] = '=IFERROR(H7/F7,"")'
    worksheet['I7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    worksheet['M7'] = '=IFERROR(L7/F7,"")'
    worksheet['M7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    worksheet['Q7'] = '=IFERROR(P7/F7,"")'
    worksheet['Q7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    # 2020
    worksheet['V7'] = '=IFERROR(U7/S7,"")'
    worksheet['V7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    worksheet['Z7'] = '=IFERROR(Y7/S7,"")'
    worksheet['Z7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    worksheet['AD7'] = '=IFERROR(AC7/S7,"")'
    worksheet['AD7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    # 2021
    worksheet['AI7'] = '=IFERROR(AH7/AF7,"")'
    worksheet['AI7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    worksheet['AM7'] = '=IFERROR(AL7/AF7,"")'
    worksheet['AM7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    worksheet['AQ7'] = '=IFERROR(AP7/AF7,"")'
    worksheet['AQ7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    # 2022
    worksheet['AV7'] = '=IFERROR(AU7/AS7,"")'
    worksheet['AV7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    worksheet['AZ7'] = '=IFERROR(AY7/AS7,"")'
    worksheet['AZ7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    worksheet['BD7'] = '=IFERROR(BC7/AS7,"")'
    worksheet['BD7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    # 2023
    worksheet['BI7'] = '=IFERROR(BH7/BF7,"")'
    worksheet['BI7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    worksheet['BM7'] = '=IFERROR(BL7/BF7,"")'
    worksheet['BM7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)
    worksheet['BQ7'] = '=IFERROR(BP7/BF7,"")'
    worksheet['BQ7'].font = openpyxl.styles.Font(name='Arial Narrow', size=14, italic=True, bold=True)

    # Нумерация строк
    row_num = 7
    for row in worksheet.iter_rows(min_row=7, min_col=2, max_col=2, values_only=True):
        if row_num == 9:
            worksheet.cell(row=row_num, column=1, value=2)
        elif row_num > 9:
            worksheet.cell(row=row_num, column=1, value=row_num - 7)
        else:
            worksheet.cell(row=row_num, column=1, value=row_num - 6)
        row_num += 1
    worksheet['A8'] = ''
    column_a = worksheet[get_column_letter(1)]
    for cell in column_a:
        cell.font = openpyxl.styles.Font(name='Arial Narrow', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    worksheet.column_dimensions['A'].width = 3.1

    # Раскраска
    fill_gray = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
    fill_orange = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')
    fill_blue = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    fill_red = Font(color='C00000', name='Arial Narrow')
    fill_green = Font(color='4F6228', name='Arial Narrow')
    fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')


    cell_ranges = [
        (9, 1000, 6, 7),  # 2019
        (3, 4, 6, 7),  # 2019
        (2, 2, 6, 17),  # 2019
        (9, 1000, 19, 20),  # 2020
        (3, 4, 19, 20),  # 2020
        (2, 2, 19, 30),  # 2020
        (9, 1000, 32, 33),  # 2021
        (3, 4, 32, 33),  # 2021
        (2, 2, 32, 43),  # 2021
        (9, 1000, 45, 46),  # 2022
        (3, 4, 45, 46),  # 2022
        (2, 2, 45, 56),  # 2022
        (9, 1000, 58, 59),  # 2023
        (3, 4, 58, 59),  # 2023
        (2, 2, 58, 69),  # 2023
    ]

    for cell_range in cell_ranges:
        for row in worksheet.iter_rows(min_row=cell_range[0], max_row=cell_range[1], min_col=cell_range[2],
                                       max_col=cell_range[3]):
            for cell in row:
                cell.fill = fill_blue

    for col in ['H', 'I', 'J', 'K', 'U', 'V', 'W', 'X', 'AH', 'AI', 'AJ', 'AK', 'AU', 'AV', 'AW', 'AX', 'BH', 'BI',
                'BJ', 'BK']:
        for cell in worksheet[col]:
            cell.font = fill_red
    for col in ['L', 'M', 'N', 'O', 'Y', 'Z', 'AA', 'AB', 'AL', 'AM', 'AN', 'AO', 'AY', 'AZ', 'BA', 'BB', 'BL', 'BM',
                'BN', 'BO']:
        for cell in worksheet[col]:
            cell.font = fill_green

    for cell in worksheet.iter_rows(min_row=7, max_row=7):
        for c in cell:
            c.fill = fill_orange
    last_row = worksheet.max_row
    for row in range(last_row, 1, -1):
        if all(cell.value is None for cell in worksheet[row]):
            last_row -= 1
        else:
            break
    for column in ['C', 'E', 'R', 'AE', 'AR', 'BE']:
        for cell in worksheet[column]:
            cell.fill = fill_gray
    for cell in worksheet[1]:
        cell.fill = fill_white

    # Удалить лишние сроки
    worksheet.delete_rows(last_row + 1, worksheet.max_row)


    no_border = Border(left=Side(), right=Side(), top=Side(), bottom=Side())
    for col in ['C', 'E', 'R', 'AE', 'AR', 'BE']:
        for cell in worksheet[col]:
            cell.border = no_border
    for row in [1]:
        for cell in worksheet[row]:
            cell.border = no_border


    # Под печать
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    worksheet.page_setup.scale = 63
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_margins = PageMargins(left=0.07, right=0.07, top=0.07, bottom=0.39, footer=0.07, header=0.07)
    worksheet.oddFooter.center.text = "Страница &P из &N"
    worksheet.page_setup.horizontalCentered = False
    worksheet.page_setup.verticalCentered = False
    worksheet.page_setup.printArea = 'A1:G50'
    worksheet.page_setup.printTitlesRow = '1:1'
    worksheet.page_setup.printTitlesCol = 'A:A'
    worksheet['D3'].alignment = Alignment(horizontal='center', vertical='center')


    workbook.save(file_path)
